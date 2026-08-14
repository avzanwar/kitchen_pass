"""Read an uploaded .csv or .xlsx into normalised rows.

The organizer's file comes from whatever they had to hand — a registration
export, a club spreadsheet, something typed on a phone. This module absorbs that
variety (encoding, delimiter, header spelling, blank padding rows) so that
`plan.py` only ever sees a clean list of dicts keyed by canonical column name.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook


class SheetError(Exception):
    """The file could not be read at all — wrong format, or no usable header."""


# ---------------------------------------------------------------------------
# Columns
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Column:
    """One canonical column, and the header spellings that map onto it."""

    key: str
    heading: str
    required: bool
    help: str
    aliases: tuple[str, ...] = ()


COLUMNS: tuple[Column, ...] = (
    Column(
        "division", "Division", True,
        "Event name. Rows sharing a name land in the same division.",
        ("divisionname", "event", "eventname", "category"),
    ),
    Column(
        "format", "Format", False,
        "doubles, singles or mixed. Default doubles.",
        ("divisionformat", "playformat", "type"),
    ),
    Column(
        "draw", "Draw", False,
        "round robin, single elim, double elim or pools. Default round robin.",
        ("drawkind", "drawformat", "drawtype", "bracket", "bracketformat"),
    ),
    Column(
        "pools", "Pools", False,
        "How many pools, when the draw is pools -> playoff. Default 2.",
        ("poolcount", "numberofpools", "nopools"),
    ),
    Column(
        "best_of", "Best of", False,
        "Games per match: 1, 3 or 5. Default 3.",
        ("bestof", "matchlength", "games", "gamespermatch"),
    ),
    Column(
        "skill", "Skill", False,
        "Optional skill bracket, e.g. 4.0.",
        ("skillbracket", "skilllevel", "level", "rating band", "ratingband"),
    ),
    Column(
        "age", "Age", False,
        "Optional age bracket, e.g. 50+.",
        ("agebracket", "agegroup"),
    ),
    Column(
        "team", "Team", False,
        "Optional team name. Left blank, it is built from the players' first names.",
        ("teamname", "pair", "pairname", "entry", "entryname"),
    ),
    Column(
        "seed", "Seed", False,
        "Optional seeding position, 1 = top seed.",
        ("seeding", "seedno", "seednumber"),
    ),
    Column(
        "player1", "Player 1", True,
        "Full name. Starts the game in the right-hand court.",
        ("p1", "playera", "playerone", "name", "playername", "player"),
    ),
    Column(
        "rating1", "Rating 1", False,
        "Optional DUPR or club rating for player 1.",
        ("r1", "dupr1", "player1rating", "rating", "duprrating"),
    ),
    Column(
        "player2", "Player 2", False,
        "Partner's full name. Required for doubles and mixed, left blank for singles.",
        ("p2", "playerb", "playertwo", "partner", "partnername"),
    ),
    Column(
        "rating2", "Rating 2", False,
        "Optional DUPR or club rating for player 2.",
        ("r2", "dupr2", "player2rating", "partnerrating"),
    ),
)

HEADINGS: tuple[str, ...] = tuple(c.heading for c in COLUMNS)

_BY_ALIAS: dict[str, str] = {}
for _column in COLUMNS:
    _BY_ALIAS[_column.key.replace("_", "")] = _column.key
    _BY_ALIAS[re.sub(r"[^a-z0-9]", "", _column.heading.lower())] = _column.key
    for _alias in _column.aliases:
        _BY_ALIAS[re.sub(r"[^a-z0-9]", "", _alias.lower())] = _column.key


def normalise_heading(raw: str) -> str:
    """Fold a header cell to its comparison key: lowercase, alphanumerics only.

    This is what lets "Player 1", "player_1", "PLAYER1" and "p1" all arrive at
    the same column without the organizer having to match a spelling exactly.
    """
    return re.sub(r"[^a-z0-9]", "", str(raw or "").lower())


@dataclass
class Sheet:
    """A parsed upload.

    `rows` are keyed by canonical column key; every key in COLUMNS is present,
    with "" for cells the file did not supply. `line` on each row is the number
    the organizer sees in their spreadsheet, so problems point somewhere real.
    """

    rows: list[dict[str, Any]] = field(default_factory=list)
    #: Canonical keys the file actually provided a column for.
    present: set[str] = field(default_factory=set)
    #: Headers we did not recognise, reported so a typo is visible rather than
    #: silently dropping a column the organizer thought they had filled in.
    unknown_headings: list[str] = field(default_factory=list)


def read_sheet(data: bytes, filename: str = "") -> Sheet:
    """Parse an uploaded file into rows. Raises `SheetError` if unreadable."""
    if not data:
        raise SheetError("The file is empty")

    table = _read_xlsx(data) if _looks_like_xlsx(data, filename) else _read_csv(data)
    return _to_rows(table)


# ---------------------------------------------------------------------------
# Format detection and raw reading
# ---------------------------------------------------------------------------


def _looks_like_xlsx(data: bytes, filename: str) -> bool:
    # .xlsx is a zip; sniff the magic rather than trusting the extension, since
    # a file renamed to .csv is a much commoner mistake than the reverse.
    if data[:2] == b"PK":
        return True
    return filename.lower().endswith((".xlsx", ".xlsm"))


def _read_xlsx(data: bytes) -> list[list[Any]]:
    try:
        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a wide range for a bad zip
        raise SheetError(
            "That .xlsx file could not be opened. If it is an older .xls, "
            "re-save it as .xlsx or CSV."
        ) from exc

    try:
        # Prefer a sheet named like the template's, so a workbook carrying
        # instructions on a second tab still imports.
        sheet = None
        for candidate in book.worksheets:
            if normalise_heading(candidate.title) in {"teams", "entries", "players"}:
                sheet = candidate
                break
        if sheet is None:
            sheet = book.worksheets[0]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        book.close()


def _read_csv(data: bytes) -> list[list[Any]]:
    text = _decode(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Sniffing fails on a single-column file, which is fine — a comma reads
        # such a file correctly anyway.
        delimiter = ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _decode(data: bytes) -> str:
    # utf-8-sig first: Excel's "CSV UTF-8" writes a BOM, and leaving it attached
    # would corrupt the very first heading and lose the Division column.
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SheetError("The file's text encoding could not be read")


# ---------------------------------------------------------------------------
# Header location and row shaping
# ---------------------------------------------------------------------------


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # openpyxl hands back numerics as floats, so a seed of 3 arrives as 3.0
        # and would otherwise be shown (and rejected) as "3.0".
        return str(int(value))
    return str(value).strip()


def _to_rows(table: list[list[Any]]) -> Sheet:
    header_index = _find_header(table)
    if header_index is None:
        raise SheetError(
            "No header row found. The first row must name the columns — "
            "at least Division and Player 1. Download the template to start from."
        )

    header = table[header_index]
    mapping: dict[int, str] = {}
    unknown: list[str] = []
    for position, raw in enumerate(header):
        text = _cell(raw)
        if not text:
            continue
        key = _BY_ALIAS.get(normalise_heading(text))
        if key is None:
            unknown.append(text)
        elif key not in mapping.values():
            mapping[position] = key

    sheet = Sheet(present=set(mapping.values()), unknown_headings=unknown)
    for offset, raw_row in enumerate(table[header_index + 1 :], start=header_index + 2):
        values = {key: _cell(raw_row[i]) if i < len(raw_row) else ""
                  for i, key in mapping.items()}
        if not any(values.values()):
            # Trailing blank rows are what spreadsheets are made of; skipping
            # them silently is right, an error per blank line is not.
            continue
        row: dict[str, Any] = {c.key: "" for c in COLUMNS}
        row.update(values)
        row["line"] = offset
        sheet.rows.append(row)

    return sheet


def _find_header(table: list[list[Any]]) -> int | None:
    """Locate the header row.

    Usually row 1, but exports often carry a title or a blank line first, so
    scan the top of the file for the row that names the most known columns.
    """
    best: tuple[int, int] | None = None
    for index, raw_row in enumerate(table[:20]):
        hits = sum(
            1
            for cell in raw_row
            if _cell(cell) and normalise_heading(_cell(cell)) in _BY_ALIAS
        )
        if hits >= 2 and (best is None or hits > best[1]):
            best = (index, hits)
    return best[0] if best else None
