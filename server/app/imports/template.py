"""The blank sheet handed to the organizer.

Generated from `sheet.COLUMNS`, so a column added to the parser appears in the
template automatically and the two cannot drift apart. The sample rows are real
enough to import as-is, which makes the template also a working demo: download,
upload, and you have a tournament.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.imports.sheet import COLUMNS, HEADINGS, Column

#: Rows shown in the template. Three divisions, deliberately covering the cases
#: an organizer needs to see: a doubles round robin, a singles knockout with
#: seeds, and a mixed pools draw — plus a repeat player across two divisions and
#: blank "as above" settings cells.
SAMPLE_ROWS: tuple[tuple[str, ...], ...] = (
    ("4.0 Men's Doubles", "doubles", "round robin", "", "3", "4.0", "",
     "Kitchen Bandits", "1", "Ivo Novak", "4.25", "Mateo Alvarez", "4.0"),
    ("4.0 Men's Doubles", "", "", "", "", "", "",
     "Dink Dynasty", "2", "Sam Whitfield", "4.0", "Toby Chen", "3.75"),
    ("4.0 Men's Doubles", "", "", "", "", "", "",
     "", "3", "Priya Raman", "4.0", "Alex Moreau", "4.25"),
    ("4.0 Men's Doubles", "", "", "", "", "", "",
     "", "4", "Chris Okafor", "3.75", "Dan Feld", "4.0"),
    ("Open Singles", "singles", "single elim", "", "3", "", "",
     "", "1", "Ivo Novak", "4.25", "", ""),
    ("Open Singles", "", "", "", "", "", "",
     "", "2", "Sam Whitfield", "4.0", "", ""),
    ("Open Singles", "", "", "", "", "", "",
     "", "3", "Toby Chen", "3.75", "", ""),
    ("Open Singles", "", "", "", "", "", "",
     "", "4", "Alex Moreau", "4.25", "", ""),
    ("50+ Mixed Doubles", "mixed", "pools", "2", "3", "", "50+",
     "", "1", "Priya Raman", "4.0", "Ivo Novak", "4.25"),
    ("50+ Mixed Doubles", "", "", "", "", "", "",
     "", "2", "Nina Roth", "3.5", "Chris Okafor", "3.75"),
    ("50+ Mixed Doubles", "", "", "", "", "", "",
     "", "3", "Grace Lim", "3.75", "Dan Feld", "4.0"),
    ("50+ Mixed Doubles", "", "", "", "", "", "",
     "", "4", "Mia Torres", "3.5", "Mateo Alvarez", "4.0"),
)

_NOTES: tuple[str, ...] = (
    "How this sheet works",
    "",
    "One row per team. Rows that share a Division name become one division.",
    "A division's settings are read from the first row that names it; leave "
    "them blank on later rows.",
    "Player names are matched against your saved roster, case-insensitively. "
    "Anyone new is added to it.",
    "Player 2 is required for doubles and mixed, and left blank for singles.",
    "Leave Team blank to name a team from its players' first names.",
    "Column order does not matter, and unknown columns are ignored.",
    "",
    "Delete these sample rows and enter your own. Upload the file to preview "
    "it before anything is created.",
)

_HEADER_FILL = PatternFill("solid", fgColor="0E7C6B")
_REQUIRED_FILL = PatternFill("solid", fgColor="0A5F52")


def template_csv() -> str:
    """The template as CSV text, including the sample rows."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(HEADINGS)
    writer.writerows(SAMPLE_ROWS)
    return buffer.getvalue()


def template_xlsx() -> bytes:
    """The template as a formatted workbook: a Teams sheet plus notes."""
    book = Workbook()
    teams: Any = book.active
    teams.title = "Teams"

    teams.append(list(HEADINGS))
    for row in SAMPLE_ROWS:
        teams.append(list(row))

    for index, column in enumerate(COLUMNS, start=1):
        cell = teams.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _REQUIRED_FILL if column.required else _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # A comment carries the per-column guidance without spending a row on
        # it, so the header stays on line 1 where every parser expects it.
        cell.comment = _comment(column)
        letter = get_column_letter(index)
        teams.column_dimensions[letter].width = max(len(column.heading) + 4, 16)

    teams.freeze_panes = "A2"
    teams.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    _add_validation(teams)

    notes = book.create_sheet("Notes")
    notes.column_dimensions["A"].width = 96
    for line in _NOTES:
        notes.append([line])
    notes["A1"].font = Font(bold=True, size=13)
    for row in notes.iter_rows(min_row=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _comment(column: Column) -> Comment:
    label = "Required. " if column.required else "Optional. "
    comment = Comment(f"{column.heading}\n\n{label}{column.help}", "Kitchen Pass")
    comment.width = 260
    comment.height = 110
    return comment


def _add_validation(sheet: Any) -> None:
    """Dropdowns on the columns with a fixed vocabulary.

    The parser accepts far more spellings than these, so validation is a
    convenience rather than the contract — hence `showErrorMessage=False`, which
    lets a pasted or typed value through instead of blocking it.
    """
    lists = {
        "format": '"doubles,singles,mixed"',
        "draw": '"round robin,single elim,double elim,pools"',
        "best_of": '"1,3,5"',
    }
    for index, column in enumerate(COLUMNS, start=1):
        formula = lists.get(column.key)
        if formula is None:
            continue
        validation = DataValidation(
            type="list", formula1=formula, allow_blank=True, showErrorMessage=False
        )
        sheet.add_data_validation(validation)
        letter = get_column_letter(index)
        validation.add(f"{letter}2:{letter}500")
