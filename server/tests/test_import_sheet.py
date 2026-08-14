"""Reading real-world spreadsheets: encodings, delimiters, header spellings.

Everything here is a shape an organizer's file has actually arrived in — an
Excel "CSV UTF-8" with a BOM, a European semicolon export, a title row above
the header, columns in the wrong order and spelled by hand.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from app.imports import SheetError, build_plan, read_sheet
from app.imports.template import template_csv, template_xlsx

BASIC = (
    "Division,Player 1,Player 2\n"
    "4.0 Mixed,Ivo Novak,Priya Raman\n"
    "4.0 Mixed,Sam Whitfield,Nina Roth\n"
)


def parse(text: str | bytes, filename: str = "teams.csv"):
    data = text.encode() if isinstance(text, str) else text
    return read_sheet(data, filename)


def test_reads_a_plain_csv() -> None:
    sheet = parse(BASIC)
    assert len(sheet.rows) == 2
    assert sheet.rows[0]["player1"] == "Ivo Novak"
    # Line numbers are what the organizer sees in their editor: header is 1.
    assert [r["line"] for r in sheet.rows] == [2, 3]


def test_strips_the_excel_utf8_bom() -> None:
    # Excel's "CSV UTF-8" prepends a BOM. Left attached it corrupts the first
    # heading, which is Division — losing the column the whole import groups on.
    sheet = parse(b"\xef\xbb\xbf" + BASIC.encode())
    assert "division" in sheet.present
    assert sheet.rows[0]["division"] == "4.0 Mixed"


def test_reads_semicolon_delimited_exports() -> None:
    sheet = parse(BASIC.replace(",", ";"))
    assert len(sheet.rows) == 2
    assert sheet.rows[1]["player2"] == "Nina Roth"


def test_reads_tab_delimited_exports() -> None:
    sheet = parse(BASIC.replace(",", "\t"))
    assert sheet.rows[0]["player1"] == "Ivo Novak"


def test_reads_latin1_when_utf8_fails() -> None:
    sheet = parse("Division,Player 1\n4.0,Ren\xe9 Dupr\xe9\n".encode("cp1252"))
    assert sheet.rows[0]["player1"] == "René Dupré"


@pytest.mark.parametrize(
    "heading",
    ["Player 1", "player_1", "PLAYER1", "p1", "Player One", "playerA", "Name"],
)
def test_header_spellings_all_reach_the_same_column(heading: str) -> None:
    sheet = parse(f"Division,{heading}\n4.0,Ivo Novak\n")
    assert sheet.rows[0]["player1"] == "Ivo Novak"


def test_column_order_does_not_matter() -> None:
    sheet = parse("Player 2,Division,Player 1\nNina Roth,4.0,Ivo Novak\n")
    assert sheet.rows[0]["division"] == "4.0"
    assert sheet.rows[0]["player1"] == "Ivo Novak"
    assert sheet.rows[0]["player2"] == "Nina Roth"


def test_skips_a_title_row_above_the_header() -> None:
    # Registration exports routinely open with a title and a blank line.
    sheet = parse("Spring Open 2026 entries\n\nDivision,Player 1\n4.0,Ivo Novak\n")
    assert len(sheet.rows) == 1
    assert sheet.rows[0]["line"] == 4


def test_ignores_blank_padding_rows() -> None:
    sheet = parse(BASIC + ",,\n,,\n\n")
    assert len(sheet.rows) == 2


def test_unknown_columns_are_reported_not_silently_dropped() -> None:
    sheet = parse(
        "Division,Player 1,Player 2,Phone,Paid\n4.0,Ivo Novak,Nina Roth,555,yes\n"
    )
    assert sheet.unknown_headings == ["Phone", "Paid"]
    plan = build_plan(sheet)
    assert plan.ok
    assert any("Phone" in p.message for p in plan.problems)


def test_a_file_with_no_header_is_rejected() -> None:
    with pytest.raises(SheetError, match="No header row"):
        parse("Ivo Novak,Priya Raman\nSam Whitfield,Nina Roth\n")


def test_an_empty_file_is_rejected() -> None:
    with pytest.raises(SheetError, match="empty"):
        parse(b"")


# ---------------------------------------------------------------------------
# xlsx
# ---------------------------------------------------------------------------


def _workbook(rows: list[list[object]], title: str = "Sheet1") -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def test_reads_an_xlsx_workbook() -> None:
    data = _workbook([["Division", "Player 1", "Player 2"],
                      ["4.0 Mixed", "Ivo Novak", "Priya Raman"]])
    sheet = read_sheet(data, "teams.xlsx")
    assert sheet.rows[0]["player2"] == "Priya Raman"


def test_xlsx_is_detected_by_content_not_extension() -> None:
    # Renaming a workbook to .csv is a far commoner mistake than the reverse.
    data = _workbook([["Division", "Player 1"], ["4.0", "Ivo Novak"]])
    sheet = read_sheet(data, "teams.csv")
    assert sheet.rows[0]["player1"] == "Ivo Novak"


def test_numeric_cells_do_not_arrive_as_floats() -> None:
    # openpyxl hands back every number as a float, so a seed of 3 would
    # otherwise be read as "3.0" and shown that way in the preview.
    data = _workbook([["Division", "Player 1", "Seed"], ["4.0", "Ivo", 3]])
    sheet = read_sheet(data, "teams.xlsx")
    assert sheet.rows[0]["seed"] == "3"


def test_prefers_a_sheet_named_teams() -> None:
    book = Workbook()
    notes = book.active
    notes.title = "Notes"
    notes.append(["How this sheet works"])
    teams = book.create_sheet("Teams")
    teams.append(["Division", "Player 1"])
    teams.append(["4.0", "Ivo Novak"])
    buffer = io.BytesIO()
    book.save(buffer)

    sheet = read_sheet(buffer.getvalue(), "book.xlsx")
    assert sheet.rows[0]["player1"] == "Ivo Novak"


def test_a_corrupt_xlsx_gives_a_useful_message() -> None:
    with pytest.raises(SheetError, match="re-save it as .xlsx or CSV"):
        read_sheet(b"PK\x03\x04 not really a workbook", "teams.xlsx")


# ---------------------------------------------------------------------------
# The template is the contract
# ---------------------------------------------------------------------------


def test_the_csv_template_imports_cleanly() -> None:
    plan = build_plan(read_sheet(template_csv().encode(), "t.csv"))
    assert plan.ok, [p.message for p in plan.problems]
    assert not plan.problems, [p.message for p in plan.problems]
    assert len(plan.divisions) == 3
    assert plan.entry_count == 12


def test_the_xlsx_template_imports_identically_to_the_csv_one() -> None:
    # The two are generated from one set of column definitions; if they ever
    # disagree, an organizer's choice of download would change their import.
    from_csv = build_plan(read_sheet(template_csv().encode(), "t.csv"))
    from_xlsx = build_plan(read_sheet(template_xlsx(), "t.xlsx"))

    def shape(plan):
        return [
            (d.name, d.format, d.draw_kind, d.best_of, d.pools,
             [(e.name, e.seed, [p.name for p in e.players]) for e in d.entries])
            for d in plan.divisions
        ]

    assert shape(from_csv) == shape(from_xlsx)


def test_the_template_covers_every_column_the_parser_accepts() -> None:
    from app.imports.sheet import COLUMNS

    sheet = read_sheet(template_csv().encode(), "t.csv")
    assert sheet.present == {c.key for c in COLUMNS}
    assert sheet.unknown_headings == []
